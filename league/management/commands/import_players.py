"""
Management command to import players from the Liberty League Fantasy Baseball Roster Excel file.
Usage: python manage.py import_players [--file path/to/roster.xlsx] [--clear]
"""
import os
from django.core.management.base import BaseCommand, CommandError
from league.models import Player, RealTeam

# Map from Excel position strings to simplified POSITION_CHOICES keys (C, IF, OF, P)
POSITION_MAP = {
    # Pitchers
    'P': 'P', 'RHP': 'P', 'LHP': 'P', 'SP': 'P', 'RP': 'P',
    # Catcher
    'C': 'C',
    # Infield (all infield positions + DH collapse to IF)
    '1B': 'IF', '2B': 'IF', '3B': 'IF', 'SS': 'IF',
    'IF': 'IF', 'INF': 'IF', 'DH': 'IF', 'UTL': 'IF',
    # Outfield
    'OF': 'OF',
    # Multi-position combos
    '1B/3B': 'IF', '1B/DH': 'IF', '1B/OF': 'IF',
    '2B/OF': 'IF', '2B/P': 'IF', '2B/SS': 'IF',
    'C/IF': 'C', 'C/OF': 'C', 'C/P': 'C', 'C/UT': 'C',
    'IF/OF': 'IF', 'IF/P': 'IF',
    'INF/OF': 'IF', 'INF/RHP': 'P',
    'LHP/OF': 'P',
    'OF/1B/C': 'OF', 'OF/C': 'OF', 'OF/IF': 'OF', 'OF/P': 'OF',
    'P/1B': 'P', 'P/IF': 'P', 'P/INF': 'P', 'P/OF': 'P',
    'RHP/IF': 'P',
}

# School abbreviations for RealTeam
SCHOOL_ABBR = {
    'Bard College': 'BARD',
    'Clarkson University': 'CU',
    'Hobart College': 'HOB',
    'Ithaca College': 'IC',
    'Rensselaer Polytechnic Institute': 'RPI',
    'Rochester Institute of Technology': 'RIT',
    'Skidmore College': 'SKI',
    'St. Lawrence University': 'SLU',
    'Union College': 'UNI',
    'University of Rochester': 'UR',
    'Vassar College': 'VAS',
}


class Command(BaseCommand):
    help = 'Import players from the Liberty League Fantasy Baseball Roster Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default='Liberty League Fantasy Baseball Roster.xlsx',
            help='Path to the Excel file (default: Liberty League Fantasy Baseball Roster.xlsx)'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Clear all existing free-agent players before importing'
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is required. Install it with: pip install openpyxl')

        filepath = options['file']
        if not os.path.isabs(filepath):
            filepath = os.path.join(os.getcwd(), filepath)

        if not os.path.exists(filepath):
            raise CommandError(f'File not found: {filepath}')

        if options['clear']:
            deleted, _ = Player.objects.filter(fantasy_team__isnull=True).delete()
            self.stdout.write(f'Cleared {deleted} existing free-agent players.')

        wb = openpyxl.load_workbook(filepath)
        if 'Draft Board' not in wb.sheetnames:
            raise CommandError("Sheet 'Draft Board' not found in the Excel file.")

        ws = wb['Draft Board']

        # Ensure all real teams exist
        real_teams = {}
        for school, abbr in SCHOOL_ABBR.items():
            # Try by name first, then by abbreviation (handles partial name matches)
            team = RealTeam.objects.filter(name=school).first()
            if team is None:
                team = RealTeam.objects.filter(abbreviation=abbr).first()
            if team is None:
                team = RealTeam.objects.create(name=school, abbreviation=abbr)
                self.stdout.write(f'  Created real team: {school}')
            else:
                # Update name to canonical form if it differs
                if team.name != school:
                    self.stdout.write(f'  Mapped "{school}" to existing team "{team.name}" ({team.abbreviation})')
            real_teams[school] = team

        imported = 0
        skipped = 0
        unknown_pos = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, class_year, position_raw, school = row[0], row[1], row[2], row[3]

            if not name or not school:
                continue

            # Skip non-players
            if position_raw == 'Manager':
                skipped += 1
                continue

            # Map position
            position = POSITION_MAP.get(str(position_raw).strip() if position_raw else '')
            if position is None:
                self.stdout.write(
                    self.style.WARNING(f'  Unknown position "{position_raw}" for {name} — skipping')
                )
                unknown_pos += 1
                continue

            real_team = real_teams.get(school)
            if real_team is None:
                self.stdout.write(
                    self.style.WARNING(f'  Unknown school "{school}" for {name} — skipping')
                )
                skipped += 1
                continue

            # Split name (handle "First Last" or "First Middle Last")
            parts = str(name).strip().split()
            if len(parts) < 2:
                first_name = parts[0]
                last_name = ''
            else:
                first_name = parts[0]
                last_name = ' '.join(parts[1:])

            class_year_val = str(class_year).strip() if class_year else ''

            # Skip duplicates
            if Player.objects.filter(
                first_name=first_name, last_name=last_name, real_team=real_team
            ).exists():
                skipped += 1
                continue

            Player.objects.create(
                first_name=first_name,
                last_name=last_name,
                position=position,
                class_year=class_year_val,
                real_team=real_team,
                fantasy_team=None,
            )
            imported += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nDone! Imported: {imported} | Skipped: {skipped} | Unknown position: {unknown_pos}'
        ))
